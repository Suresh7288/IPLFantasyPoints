from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd
import logging
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)

# Predefined teams
TEAMS = {
    "Team Anirudh": [
        "Tilak Varma", "Noor Ahmad", "Ishan Kishan", "Pat Cummins", "Sunil Narine",
        "Heinrich Klaasen", "Rashid Khan", "Krunal Pandya", "Ravindra Jadeja", "Shivam Dube",
        "Mitchell Santner", "Glenn Phillips", "Wanindu Hasaranga", "Ajinkya Rahane", "Nitish Kumar R"
    ],
    "Team Bhanu": [
        "Mohammed Siraj", "Josh Hazlewood", "Mitchell Marsh", "Shardul Thakur", "Andre Russell",
        "Liam Livingstone", "Rishabh Pant", "Shreyas Iyer", "Ashwani", "Quinton de Kock",
        "Priyansh Arya", "Jitesh Sharma", "Marcus Stoinis", "Aniket Verma", "Vipraj",
        "Ryan Rickelton", "Sai Kishore", "Kagiso Rabada"
    ],
    "Team Satya": [
        "Trent Boult", "Abhishek Sharma", "Raghuvanshi", "Glenn Maxwell", "Yashasvi Jaiswal",
        "Phil Salt", "Shubman Gill", "Hardik Pandya", "Kuldeep Yadav", "Jasprit Bumrah",
        "Mitchell Starc", "Sanju Samson", "Jos Buttler", "Tristan Stubbs", "Prabhsimran Singh",
        "Jofra Archer", "Sherfane Rutherford", "Prince Yadav"
    ],
    "Team Sunny": [
        "Rajat Patidar", "Tim David", "Riyan Parag", "Nicholas Pooran", "KL Rahul",
        "Sai Sudharsan", "Mohammed Shami", "Ravichandran Ashwin", "Moeen Ali", "Jake",
        "Harshal Patel", "Axar Patel", "Venkatesh Iyer", "Omarzai", "Maheesh Theekshana",
        "Simarjeet Singh", "Sameer Rizvi", "Vignesh Puth"
    ],
    "Team Suresh": [
        "Nitish Rana", "Ravi Bishnoi", "Virat Kohli", "Travis Head", "Matheesha Pathirana",
        "Shahrukh Khan", "Varun Chakravarthy", "Yash Dayal", "Rahul Tripathi", "Ayush Badoni",
        "Faf du Plessis", "Aiden Markram", "Rinku Singh", "Dhruv Jurel", "Marco Jansen",
        "Adam Zampa", "Arshdeep Singh", "Porel"
    ],
    "Team Trivedh": [
        "Rachin Ravindra", "Ruturaj Gaikwad", "Will Jacks", "MS Dhoni", "Shahbaz Ahmed",
        "Suryakumar Yadav", "Harshit Rana", "Suyash Sharma", "David Miller", "Rohit Sharma",
        "Khaleel Ahmed", "Yuzvendra Chahal", "Ashutosh Sharma", "Prasidh Krishna", "Bhuvneshwar Kumar",
        "Devdutt Padikkal", "Deepak Chahar", "Shimron Hetmyer"
    ]
}
FIXED_EXCEL_FILE = "ipl_player_stats.xlsx"


def initialize_driver():
    """Initialize and return Chrome WebDriver."""
    try:
        logging.info("Initializing browser...")
        return webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    except Exception as e:
        logging.error(f"Failed to initialize browser: {str(e)}")
        return None


def login_to_ipl_fantasy(driver):
    """Log in to IPL Fantasy website."""
    try:
        driver.get("https://fantasy.iplt20.com/my11c/static/login.html")
        time.sleep(2)

        email = input("Enter your email: ")
        driver.find_element(By.ID, "email_input").send_keys(email)
        driver.find_element(By.ID, "registerCTA").click()
        logging.info("OTP sent to your email")
        time.sleep(2)

        otp = input("Enter the 6-digit OTP: ")
        driver.find_element(By.ID, "otpInputField").send_keys(otp)
        driver.find_element(By.ID, "verifyOtp").click()
        logging.info("Login successful!")

        WebDriverWait(driver, 20).until(lambda d: "classic" in d.current_url)

        stats_url = "https://fantasy.iplt20.com/classic/stats"
        if "classic/stats" not in driver.current_url:
            logging.info("Redirecting to stats page...")
            driver.get(stats_url)

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "li > div.m11c-tbl"))
        )
        return True

    except Exception as e:
        logging.error(f"Login failed: {str(e)}")
        return False


def fetch_player_stats(driver, player_name):
    """Fetch stats for a single player."""
    try:
        players = driver.find_elements(By.CSS_SELECTOR,
                                       "li > div.m11c-tbl > div.m11c-tbl__body > div.m11c-tbl__row")

        for player in players:
            try:
                name = player.find_element(
                    By.CSS_SELECTOR, ".m11c-tbl__cell--name .m11c-plyrSel__name span"
                ).text.strip()

                if player_name.lower() in name.lower():
                    team = player.find_element(
                        By.CSS_SELECTOR, ".m11c-tbl__cell--thumb .m11c-plyrSel__team span"
                    ).text.strip()
                    points = player.find_element(
                        By.CSS_SELECTOR, ".m11c-tbl__cell--amt span"
                    ).text.strip()

                    # Convert points to integer (handle commas if present)
                    try:
                        points = int(points.replace(',', ''))
                    except:
                        points = 0

                    return {
                        'Player Name': name,
                        'IPL Team': team,
                        'Points': points
                    }
            except Exception:
                continue

        logging.warning(f"Player '{player_name}' not found")
        return None

    except Exception as e:
        logging.error(f"Error searching for player {player_name}: {str(e)}")
        return None


def process_teams(driver, teams):
    """Process all teams and return collected data with team totals."""
    all_team_data = {}

    for team_name, players in teams.items():
        logging.info(f"\nProcessing {team_name}...")
        driver.refresh()
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "li > div.m11c-tbl > div.m11c-tbl__body"))
        )

        team_players = []
        for player in players:
            player_data = fetch_player_stats(driver, player)
            if player_data:
                team_players.append(player_data)

        # Calculate team total (sum of top 11 players)
        sorted_players = sorted(team_players, key=lambda x: x['Points'], reverse=True)
        top_players = sorted_players[:11]  # Get top 11 players
        team_total = sum(player['Points'] for player in top_players)

        all_team_data[team_name] = {
            'players': team_players,
            'total_points': team_total
        }

    return all_team_data


def create_excel_with_team_format(team_data):
    """Create Excel file with team-wise formatting and totals."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Player Stats"

        # Styles
        header_font = Font(bold=True, size=12)
        team_header_font = Font(bold=True, size=14, color="FFFFFF")
        total_font = Font(bold=True, size=12, color="FFFFFF")
        team_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green color
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        row_num = 1

        for team_name, data in team_data.items():
            players = data['players']
            team_total = data['total_points']

            # Add team header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            cell = ws.cell(row=row_num, column=1, value=team_name)
            cell.font = team_header_font
            cell.fill = team_header_fill
            cell.alignment = Alignment(horizontal='center')
            row_num += 1

            # Add column headers
            headers = ['Player Name', 'IPL Team', 'Points', 'In Top 11']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.border = border

            row_num += 1

            # Sort players by points descending
            sorted_players = sorted(players, key=lambda x: x['Points'], reverse=True)

            # Add player data
            for i, player in enumerate(sorted_players, 1):
                ws.cell(row=row_num, column=1, value=player['Player Name']).border = border
                ws.cell(row=row_num, column=2, value=player['IPL Team']).border = border
                ws.cell(row=row_num, column=3, value=player['Points']).border = border

                # Mark if player is in top 11
                in_top_11 = "Yes" if i <= 11 else "No"
                ws.cell(row=row_num, column=4, value=in_top_11).border = border

                row_num += 1

            # Add team total row
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
            cell = ws.cell(row=row_num, column=1, value=f"Total Points (Top 11 Players)")
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal='right')

            cell = ws.cell(row=row_num, column=4, value=team_total)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border

            row_num += 2  # Add extra space between teams

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12

        # Save the file
        wb.save(FIXED_EXCEL_FILE)
        logging.info(f"Data saved to {os.path.abspath(FIXED_EXCEL_FILE)}")
        return True

    except Exception as e:
        logging.error(f"Failed to create Excel file: {str(e)}")
        return False


def main():
    print("IPL Fantasy Team Points Tracker")
    print("=" * 40)

    # Initialize browser
    driver = initialize_driver()
    if not driver:
        return

    try:
        # Login
        if not login_to_ipl_fantasy(driver):
            return

        # Process all teams
        team_data = process_teams(driver, TEAMS)

        # Create formatted Excel file
        if create_excel_with_team_format(team_data):
            print("\nSuccessfully saved data with team totals!")

            # # Print summary
            # for team_name, data in team_data.items():
            #     print(f"\n{team_name} (Total: {data['total_points']} pts):")
            #     for i, player in enumerate(sorted(data['players'], key=lambda x: x['Points'], reverse=True), 1):
            #         top_marker = "*" if i <= 11 else ""
            #         print(f"- {player['Player Name']}: {player['Points']} pts{top_marker}")

        else:
            print("\nFailed to save data. Check log for details.")

    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
    finally:
        driver.quit()
        print("\nSession ended.")


if __name__ == "__main__":
    main()